"""Excel export: セッションstateのプロジェクトデータを.xlsxに書き出す"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .templates import (
    AI_QUERY_TEMPLATES, AI_SERVICES, SCHEMA_ITEMS, EEAT_ITEMS,
    TECH_SEO_ITEMS, AREA_ORDER, AREAS,
)
from .scoring import score_all


HEADER_FILL = PatternFill("solid", start_color="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
INPUT_FILL = PatternFill("solid", start_color="FFFBEB")
CALC_FILL = PatternFill("solid", start_color="F1F5F9")
TOTAL_FILL = PatternFill("solid", start_color="DBEAFE")
SECTION_FILL = PatternFill("solid", start_color="E0E7FF")
TOTAL_FONT = Font(bold=True, size=11, name="Arial")
TITLE_FONT = Font(bold=True, size=18, color="1E293B", name="Arial")
REGULAR_FONT = Font(size=10, name="Arial")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="94A3B8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _sh(cell, fill=None, font=None, align=CENTER, border=True):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    cell.alignment = align
    if border:
        cell.border = BORDER


def build_xlsx(project: dict) -> bytes:
    """プロジェクトdictから.xlsxバイト列を生成"""
    wb = Workbook()

    # ======== Cover ========
    ws = wb.active
    ws.title = "Cover"

    ws["A1"] = "AIO/SEO診断 結果"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"出力日時: {project.get('updated_at', '')} | AIO-Scope v1"
    ws["A2"].font = Font(italic=True, size=9, color="64748B", name="Arial")
    ws.merge_cells("A2:E2")

    # プロジェクト情報
    ws["A4"] = "プロジェクト情報"
    _sh(ws["A4"], fill=SECTION_FILL, font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    ws.merge_cells("A4:E4")

    info_rows = [
        ("案件名", project.get("name", "")),
        ("クライアント名", project.get("client_name", "")),
        ("対象URL", project.get("target_url", "")),
        ("業界カテゴリ", project.get("industry", "")),
        ("担当者", project.get("owner_email", "")),
        ("競合社", ", ".join(project.get("competitors", []))),
    ]
    for i, (k, v) in enumerate(info_rows, start=5):
        cell_k = ws.cell(row=i, column=1, value=k)
        _sh(cell_k, fill=HEADER_FILL, font=HEADER_FONT)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
        cell_v = ws.cell(row=i, column=2, value=v)
        _sh(cell_v, fill=INPUT_FILL, font=REGULAR_FONT, align=LEFT)

    # スコアサマリー
    sc = score_all(project)
    score_start = 12
    cell_sh = ws.cell(row=score_start, column=1, value="スコアサマリー")
    _sh(cell_sh, fill=SECTION_FILL, font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    ws.merge_cells(start_row=score_start, start_column=1, end_row=score_start, end_column=5)

    headers = ["#", "領域", "満点", "スコア", "充足率"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=score_start + 1, column=i, value=h)
        _sh(c, fill=HEADER_FILL, font=HEADER_FONT)

    for i, key in enumerate(AREA_ORDER):
        a = AREAS[key]
        r = score_start + 2 + i
        c1 = ws.cell(row=r, column=1, value=f"領域{a['num']}")
        _sh(c1, fill=HEADER_FILL, font=HEADER_FONT)
        c2 = ws.cell(row=r, column=2, value=a["name"])
        _sh(c2, fill=INPUT_FILL, font=REGULAR_FONT, align=LEFT)
        c3 = ws.cell(row=r, column=3, value=a["max_points"])
        _sh(c3, fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))
        c4 = ws.cell(row=r, column=4, value=sc[key])
        _sh(c4, fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))
        pct = sc[key] / a["max_points"] if a["max_points"] > 0 else 0
        c5 = ws.cell(row=r, column=5, value=pct)
        c5.number_format = "0.0%"
        _sh(c5, fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))

    total_r = score_start + 2 + len(AREA_ORDER)
    ct1 = ws.cell(row=total_r, column=1, value="合計")
    _sh(ct1, fill=TOTAL_FILL, font=TOTAL_FONT)
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=2)
    ct3 = ws.cell(row=total_r, column=3, value=100)
    _sh(ct3, fill=TOTAL_FILL, font=TOTAL_FONT)
    ct4 = ws.cell(row=total_r, column=4, value=sc["total"])
    _sh(ct4, fill=TOTAL_FILL, font=TOTAL_FONT)
    ct5 = ws.cell(row=total_r, column=5, value=sc["total"] / 100)
    ct5.number_format = "0.0%"
    _sh(ct5, fill=TOTAL_FILL, font=TOTAL_FONT)

    # ランク
    from .templates import rank_from_score
    rank, desc, _ = rank_from_score(sc["total"])
    rank_r = total_r + 2
    cr1 = ws.cell(row=rank_r, column=1, value="総合ランク")
    _sh(cr1, fill=SECTION_FILL, font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    ws.merge_cells(start_row=rank_r, start_column=1, end_row=rank_r, end_column=2)
    cr3 = ws.cell(row=rank_r, column=3, value=rank)
    cr3.font = Font(bold=True, size=20, color="1E293B", name="Arial")
    cr3.fill = TOTAL_FILL
    cr3.alignment = CENTER
    cr3.border = BORDER
    cr4 = ws.cell(row=rank_r, column=4, value=desc)
    _sh(cr4, fill=TOTAL_FILL, font=REGULAR_FONT, align=LEFT)
    ws.merge_cells(start_row=rank_r, start_column=4, end_row=rank_r, end_column=5)

    for col, w in [("A", 10), ("B", 28), ("C", 10), ("D", 12), ("E", 14)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # ======== Area 1: AI引用 ========
    ws2 = wb.create_sheet("01_AI引用")
    ws2["A1"] = f"領域1: AI引用露出度 (スコア: {sc['ai_quote']}/20)"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:Q1")
    ws2.row_dimensions[1].height = 24

    ws2.cell(row=2, column=1, value="No")
    ws2.cell(row=2, column=2, value="種別")
    ws2.cell(row=2, column=3, value="クエリ")
    ws2.merge_cells("A2:A3")
    ws2.merge_cells("B2:B3")
    ws2.merge_cells("C2:C3")
    col_start = 4
    for ai in AI_SERVICES:
        ws2.cell(row=2, column=col_start, value=ai)
        ws2.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + 2)
        ws2.cell(row=3, column=col_start, value="引用")
        ws2.cell(row=3, column=col_start + 1, value="位置")
        ws2.cell(row=3, column=col_start + 2, value="正確性")
        col_start += 3
    ws2.cell(row=2, column=16, value="小計")
    ws2.merge_cells("P2:P3")
    ws2.cell(row=2, column=17, value="メモ")
    ws2.merge_cells("Q2:Q3")
    for c in range(1, 18):
        _sh(ws2.cell(row=2, column=c), fill=HEADER_FILL, font=HEADER_FONT)
        _sh(ws2.cell(row=3, column=c), fill=HEADER_FILL, font=HEADER_FONT)

    rows = project.get("diagnosis", {}).get("ai_quote", {}).get("rows", [])
    for i, row in enumerate(rows):
        r = 4 + i
        _sh(ws2.cell(row=r, column=1, value=i + 1), fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))
        _sh(ws2.cell(row=r, column=2, value=row.get("category", "")), fill=INPUT_FILL, font=REGULAR_FONT)
        _sh(ws2.cell(row=r, column=3, value=row.get("query", row.get("template", ""))),
            fill=INPUT_FILL, font=REGULAR_FONT, align=LEFT)
        subtotal = 0
        ci = 4
        for ai in AI_SERVICES:
            s = row.get(ai, {})
            cite = int(s.get("cite", 0) or 0)
            pos = int(s.get("position", 0) or 0)
            acc = int(s.get("accuracy", 0) or 0)
            _sh(ws2.cell(row=r, column=ci, value=cite), fill=INPUT_FILL, font=REGULAR_FONT)
            _sh(ws2.cell(row=r, column=ci + 1, value=pos), fill=INPUT_FILL, font=REGULAR_FONT)
            _sh(ws2.cell(row=r, column=ci + 2, value=acc), fill=INPUT_FILL, font=REGULAR_FONT)
            subtotal += cite + pos + acc
            ci += 3
        _sh(ws2.cell(row=r, column=16, value=subtotal), fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))
        _sh(ws2.cell(row=r, column=17, value=""), fill=INPUT_FILL, font=REGULAR_FONT, align=LEFT)

    widths = {"A": 5, "B": 12, "C": 28}
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w
    for col in "DEFGHIJKLMNO":
        ws2.column_dimensions[col].width = 8
    ws2.column_dimensions["P"].width = 10
    ws2.column_dimensions["Q"].width = 20
    ws2.freeze_panes = "D4"

    # ======== Area 2: SGE ========
    ws3 = wb.create_sheet("02_SGE")
    ws3["A1"] = f"領域2: Google SGE (スコア: {sc['sge']}/15)"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:G1")

    hdrs = ["No", "キーワード", "検索Vol", "AI Overview", "自社引用", "リンク", "メモ"]
    for i, h in enumerate(hdrs, start=1):
        _sh(ws3.cell(row=2, column=i, value=h), fill=HEADER_FILL, font=HEADER_FONT)

    sge_rows = project.get("diagnosis", {}).get("sge", {}).get("rows", [])
    for i, row in enumerate(sge_rows):
        r = 3 + i
        _sh(ws3.cell(row=r, column=1, value=i + 1), fill=CALC_FILL)
        _sh(ws3.cell(row=r, column=2, value=row.get("keyword", "")), fill=INPUT_FILL, align=LEFT)
        _sh(ws3.cell(row=r, column=3, value=row.get("volume", "")), fill=INPUT_FILL)
        _sh(ws3.cell(row=r, column=4, value="○" if row.get("overview") else ""), fill=INPUT_FILL)
        _sh(ws3.cell(row=r, column=5, value="○" if row.get("cite") else ""), fill=INPUT_FILL)
        _sh(ws3.cell(row=r, column=6, value="○" if row.get("link") else ""), fill=INPUT_FILL)
        _sh(ws3.cell(row=r, column=7, value=row.get("note", "")), fill=INPUT_FILL, align=LEFT)
    for col, w in [("A", 5), ("B", 30), ("C", 10), ("D", 12), ("E", 10), ("F", 8), ("G", 22)]:
        ws3.column_dimensions[col].width = w

    # ======== Area 3: Schema ========
    ws4 = wb.create_sheet("03_構造化データ")
    ws4["A1"] = f"領域3: 構造化データ (スコア: {sc['schema']}/20)"
    ws4["A1"].font = TITLE_FONT
    ws4.merge_cells("A1:E1")

    for i, h in enumerate(["No", "項目", "配点", "実装(0-2)", "取得"]):
        _sh(ws4.cell(row=2, column=i + 1, value=h), fill=HEADER_FILL, font=HEADER_FONT)
    items = project.get("diagnosis", {}).get("schema", {}).get("items", {})
    for i, (name, pt) in enumerate(SCHEMA_ITEMS):
        r = 3 + i
        impl = int(items.get(str(i), 0) or 0)
        _sh(ws4.cell(row=r, column=1, value=i + 1), fill=CALC_FILL)
        _sh(ws4.cell(row=r, column=2, value=name), fill=INPUT_FILL, align=LEFT)
        _sh(ws4.cell(row=r, column=3, value=pt), fill=CALC_FILL)
        _sh(ws4.cell(row=r, column=4, value=impl), fill=INPUT_FILL)
        _sh(ws4.cell(row=r, column=5, value=round(pt * impl / 2, 2)), fill=CALC_FILL)
    for col, w in [("A", 5), ("B", 40), ("C", 8), ("D", 12), ("E", 10)]:
        ws4.column_dimensions[col].width = w

    # ======== Area 4: E-E-A-T ========
    ws5 = wb.create_sheet("04_EEAT")
    ws5["A1"] = f"領域4: E-E-A-T (スコア: {sc['eeat']}/15)"
    ws5["A1"].font = TITLE_FONT
    ws5.merge_cells("A1:D1")
    for i, h in enumerate(["No", "項目", "配点", "取得"]):
        _sh(ws5.cell(row=2, column=i + 1, value=h), fill=HEADER_FILL, font=HEADER_FONT)
    eeat_items = project.get("diagnosis", {}).get("eeat", {}).get("items", {})
    for i, (name, pt) in enumerate(EEAT_ITEMS):
        r = 3 + i
        val = float(eeat_items.get(str(i), 0) or 0)
        _sh(ws5.cell(row=r, column=1, value=i + 1), fill=CALC_FILL)
        _sh(ws5.cell(row=r, column=2, value=name), fill=INPUT_FILL, align=LEFT)
        _sh(ws5.cell(row=r, column=3, value=pt), fill=CALC_FILL)
        _sh(ws5.cell(row=r, column=4, value=val), fill=INPUT_FILL)
    for col, w in [("A", 5), ("B", 40), ("C", 8), ("D", 10)]:
        ws5.column_dimensions[col].width = w

    # ======== Area 5: Tech SEO ========
    ws6 = wb.create_sheet("05_テクニカルSEO")
    ws6["A1"] = f"領域5: テクニカルSEO (スコア: {sc['tech_seo']}/15)"
    ws6["A1"].font = TITLE_FONT
    ws6.merge_cells("A1:D1")
    for i, h in enumerate(["No", "項目", "配点", "取得"]):
        _sh(ws6.cell(row=2, column=i + 1, value=h), fill=HEADER_FILL, font=HEADER_FONT)
    tech_items = project.get("diagnosis", {}).get("tech_seo", {}).get("items", {})
    for i, (name, pt) in enumerate(TECH_SEO_ITEMS):
        r = 3 + i
        val = float(tech_items.get(str(i), 0) or 0)
        _sh(ws6.cell(row=r, column=1, value=i + 1), fill=CALC_FILL)
        _sh(ws6.cell(row=r, column=2, value=name), fill=INPUT_FILL, align=LEFT)
        _sh(ws6.cell(row=r, column=3, value=pt), fill=CALC_FILL)
        _sh(ws6.cell(row=r, column=4, value=val), fill=INPUT_FILL)
    for col, w in [("A", 5), ("B", 40), ("C", 8), ("D", 10)]:
        ws6.column_dimensions[col].width = w

    # ======== Area 6: Competitor ========
    ws7 = wb.create_sheet("06_競合")
    ws7["A1"] = f"領域6: 競合ベンチマーク (スコア: {sc['competitor']}/15)"
    ws7["A1"].font = TITLE_FONT
    ws7.merge_cells("A1:G1")
    hdrs7 = ["会社名", "領域1", "領域2", "領域3", "領域4", "領域5", "合計"]
    for i, h in enumerate(hdrs7):
        _sh(ws7.cell(row=2, column=i + 1, value=h), fill=HEADER_FILL, font=HEADER_FONT)
    # 自社
    _sh(ws7.cell(row=3, column=1, value=f"{project.get('client_name', '自社')} (自社)"),
        fill=TOTAL_FILL, font=TOTAL_FONT, align=LEFT)
    self_scores = [sc["ai_quote"], sc["sge"], sc["schema"], sc["eeat"], sc["tech_seo"]]
    for i, s in enumerate(self_scores):
        _sh(ws7.cell(row=3, column=2 + i, value=s), fill=TOTAL_FILL, font=TOTAL_FONT)
    _sh(ws7.cell(row=3, column=7, value=round(sum(self_scores), 1)), fill=TOTAL_FILL, font=TOTAL_FONT)
    # 競合
    comps = project.get("diagnosis", {}).get("competitor", {}).get("competitors", [])
    for i, c in enumerate(comps):
        r = 4 + i
        _sh(ws7.cell(row=r, column=1, value=c.get("name", "")), fill=INPUT_FILL, align=LEFT)
        scores = c.get("scores", [0, 0, 0, 0, 0])
        for j, s in enumerate(scores):
            _sh(ws7.cell(row=r, column=2 + j, value=float(s or 0)), fill=INPUT_FILL)
        _sh(ws7.cell(row=r, column=7, value=round(sum(float(s or 0) for s in scores), 1)),
            fill=CALC_FILL, font=Font(bold=True, size=10, name="Arial"))
    for col, w in [("A", 28), ("B", 10), ("C", 10), ("D", 10), ("E", 10), ("F", 10), ("G", 10)]:
        ws7.column_dimensions[col].width = w

    # ======== Findings / Actions (optional sheets) ========
    ws8 = wb.create_sheet("所見・改善提案")
    ws8["A1"] = "所見・改善提案"
    ws8["A1"].font = TITLE_FONT
    ws8.merge_cells("A1:C1")

    ws8["A3"] = "強み"
    _sh(ws8["A3"], fill=SECTION_FILL, font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    strengths = project.get("findings", {}).get("strengths", [])
    for i, s in enumerate(strengths):
        ws8.cell(row=4 + i, column=1, value=f"・{s}").alignment = LEFT

    start = 4 + max(len(strengths), 1) + 2
    ws8.cell(row=start, column=1, value="弱み")
    _sh(ws8.cell(row=start, column=1), fill=SECTION_FILL,
        font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    weaknesses = project.get("findings", {}).get("weaknesses", [])
    for i, w in enumerate(weaknesses):
        ws8.cell(row=start + 1 + i, column=1, value=f"・{w}").alignment = LEFT

    start2 = start + 1 + max(len(weaknesses), 1) + 2
    ws8.cell(row=start2, column=1, value="改善アクション")
    _sh(ws8.cell(row=start2, column=1), fill=SECTION_FILL,
        font=Font(bold=True, size=11, name="Arial"), align=LEFT)
    ws8.merge_cells(start_row=start2, start_column=1, end_row=start2, end_column=3)
    for i, a in enumerate(project.get("actions", [])):
        r = start2 + 1 + i
        ws8.cell(row=r, column=1, value=a.get("title", "")).alignment = LEFT
        ws8.cell(row=r, column=2, value=a.get("impact", "")).alignment = LEFT
        ws8.cell(row=r, column=3, value=a.get("effort", "")).alignment = LEFT

    ws8.column_dimensions["A"].width = 45
    ws8.column_dimensions["B"].width = 12
    ws8.column_dimensions["C"].width = 14

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
